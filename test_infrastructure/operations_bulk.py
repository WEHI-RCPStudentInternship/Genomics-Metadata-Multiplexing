import os
import pandas as pd
import openpyxl
import fcsparser

# standard well dimensions
WELL_ROWS = 16
WELL_COLS = 24 

# how far we search for the start of the well positions and sample lookup
ROWMAX = 100
COLMAX = 25

def get_well_positions(meta):
    '''
    Extract well positions (e.g., A1, A2 etc.) from FCS file metadata,
    ordered by sorting locations.   
    '''
    metalist = list(meta.keys())
    locnames = [item for item in metalist if item.startswith('INDEX SORTING LOCATIONS')]
    locnames = sorted(locnames, key=lambda s: int(s.split('_')[1]))

    wells = []
    for locname in locnames:
        loclist = meta[locname]
        locs = loclist.split(";")

        for loc in locs:
            if loc:
                x, y = loc.split(",")
                new_x = chr(int(x) + ord('A'))
                new_y = str(int(y) + 1)
                wells.append(new_x + new_y)

    return wells

def get_plate_and_sample_from_filepath(fcs_filepath):
    '''
    Extract plate and sample name from file names in format
    e.g., DDmonthYY_INX_samplename_platename.fcs where plate
    name starts with LCE.
    '''
    filename = os.path.basename(fcs_filepath)
    filename = os.path.splitext(filename)[0]

    # Extract the plate name
    plate_start_index = filename.find('LCE')
    if plate_start_index != -1:
        plate = filename[plate_start_index:]

    # Extract the sample name
    sample_start_index = filename.find('INX_')
    if sample_start_index != -1:
        sample_name = filename[sample_start_index + len('INX_'):(plate_start_index - 1)]

    return plate, sample_name

def collate_fcs_files(fcs_files, upload_dir):
    '''
    Collate FCS files into a single dataframe, adding columns for plate and sample name.
    '''
    fcs_data = pd.DataFrame()

    for fcs_file in fcs_files:
        # fcs_savepath = os.path.join(upload_dir, fcs_file.filename)
        fcs_savepath = fcs_file  # alter this function for mock up test result purpose
        meta, data = fcsparser.parse(fcs_savepath, meta_data_only=False, reformat_meta=True)
        data = data.sort_values('Time')
        
        plate, sample = get_plate_and_sample_from_filepath(fcs_savepath)
        data['Plate#'] = plate
        data['Well position'] = get_well_positions(meta)
        data['Sample name'] = sample

        fcs_data = pd.concat([fcs_data, data])

    return fcs_data

def get_sample_lookup(sheet, sample_start_cell):
    '''
    Extract sample lookup (colour index > sample name) from plate layout spreadsheet.
    '''
    sample_dict = {}

    current_cell = sample_start_cell
    tolerate_n_blank_rows = 2
    while(True):
        current_cell = sheet.cell(row=current_cell.row + 1, column=current_cell.column)
        colour = sheet.cell(row=current_cell.row, column=current_cell.column).fill.start_color.index

        if colour == '00000000':
            if tolerate_n_blank_rows == 0:
                break
            tolerate_n_blank_rows -= 1
            continue

        value = sheet.cell(current_cell.row, current_cell.column + 1).value
        sample_dict[colour] = value

    return sample_dict

def get_sample_list(sheet, sample_lookup, well_start_cell):
    '''
    iterate through the plate cells and extract
    sample name for each cell from the colour
    '''
    sample_list = []
    for row in range(0, WELL_ROWS):
        for col in range(1, WELL_COLS + 1):
            cell = sheet.cell(row=well_start_cell.row + row, column=well_start_cell.column + col)

            samplename = 'removed'
            if cell.fill.start_color.index != '00000000':
                cell_colour = cell.fill.start_color.index

                error_message = 'Cell colour {} in cell {} not found in sample lookup in sheet {}'.format(
                    cell_colour, cell.coordinate, sheet.title
                )
                assert cell_colour in sample_lookup, error_message
                samplename = sample_lookup[cell_colour]
            
            # get well ID
            well_row = sheet.cell(row=well_start_cell.row + row, column=well_start_cell.column).value
            well_col = sheet.cell(row=well_start_cell.row - 1, column=well_start_cell.column + col).value
            well_id = well_row + str(well_col)

            sample_list.append((well_id, samplename))

    return sample_list

def find_sample_start_cell(sheet):
    '''
    Find the start of the sample names in the spreadsheet.
    Return the cell.
    '''
    sample_start = None

    for col in range(1, COLMAX):
        for row in range(1, ROWMAX):
            cell = sheet.cell(row=row, column=col)
            if cell.value and str(cell.value).lower() in ['sort description', 'sort discription']:
                sample_start = cell
                break

    return sample_start

def find_well_start(sheet):
    '''
    Find the start of the well positions in the spreadsheet.
    This is the first cell containing just the value 'A' with
    a cell containing '1' above and to the right.
    '''
    well_start_cell = None
    for col in range(1, COLMAX):
        for row in range(1, ROWMAX):
            cell = sheet.cell(row=row, column=col)
            if cell.value == 'A':
                well_start_cell = cell
                # check that the cell above and to the right is 1
                if sheet.cell(row=row - 1, column=col + 1).value == 1:
                    return well_start_cell
    return None

def plate_to_samplesheet(xlsx_file):
    '''
    Convert plate layout spreadsheet to samplesheet.
    '''
    wb = openpyxl.load_workbook(xlsx_file)

    # iterate through all sheets
    full_samplesheet = pd.DataFrame()
    for sheet in wb:
        if not sheet.title.lower().startswith('lce'):
            print('Skipping sheet {}'.format(sheet.title))
            continue

        sample_start_cell = find_sample_start_cell(sheet)
        assert sample_start_cell, 'Could not find sample start cell in sheet {}'.format(sheet.title)

        well_start_cell = find_well_start(sheet)
        assert well_start_cell, 'Could not find well start cell in sheet {}'.format(sheet.title)

        sample_lookup = get_sample_lookup(sheet, sample_start_cell)
        sample_list = get_sample_list(sheet, sample_lookup, well_start_cell)
        samplesheet = pd.DataFrame(sample_list, columns=['well_position', 'sample'])
        samplesheet['plate'] = sheet.title

        full_samplesheet = pd.concat([full_samplesheet, samplesheet])

    # reorder columns
    full_samplesheet = full_samplesheet[['plate', 'well_position', 'sample']]
    full_samplesheet.rename({'plate': 'Plate#', 'well_position': 'Well position', 'sample': "Sample name"}, axis=1, inplace=True)
    return full_samplesheet

def load_excel_samplesheet(template_sheet_filepath):
    df = pd.read_excel(template_sheet_filepath, skiprows=1, header=None)

    def combine_columns(a, b):
        if pd.isna(a):
            return b
        elif pd.isna(b):
            return a
        return f'{a}_{b}'
        
    new_columns = [combine_columns(a, b) for a, b in zip(df.iloc[0], df.iloc[1])]

    df.columns = new_columns
    df = df.drop(index=[0, 1])

    # df.rename({'Plate#': 'plate', 'Well position': 'well_position', 'Sample name': 'sample'}, axis=1, inplace=True)    
    return df


def merge_data_with_samplesheet(spreadsheet_filepath, fcs_file, template_sheet_filepath):
    '''
    Merges processed/uploaded data (may be sample sheet and/or fcs data).
    Will merge into a template if provided.
    '''
    fcs_data = pd.read_csv(fcs_file, sep='\t') if fcs_file else None
    is_xlsx = spreadsheet_filepath.endswith('.xlsx')
    merged_data = pd.DataFrame()

    if template_sheet_filepath and is_xlsx:
        raise Exception('Cannot merge template sheet with xlsx file. Please upload a tsv file with plate, sample and well positions.')
    elif template_sheet_filepath:
       
        template = load_excel_samplesheet(template_sheet_filepath)
        spreadsheet = pd.read_csv(spreadsheet_filepath, sep='\t')  # sample sheet
        
        for plate in spreadsheet['Plate#'].unique():
            
            plate_data = template.copy()
            plate_data.rename({'Sample name': 'Sample name'}, axis=1, inplace=True)
            plate_data['Plate#'] = plate
            # print(plate_data.columns, spreadsheet.columns)
            plate_data = pd.merge(plate_data, spreadsheet,
                                  on=['Plate#', 'Well position', 'Sample name'], how='left')
            # return plate_data
            merged_data = pd.concat([merged_data, plate_data])
        
    elif not is_xlsx:
        merged_data = pd.read_csv(spreadsheet_filepath, sep='\t')

    if fcs_data is None:
        return merged_data
    
    if is_xlsx:
        spreadsheet = load_excel_samplesheet(spreadsheet_filepath)
        samples_colname = [col for col in spreadsheet.columns if col.lower() == 'sample' or col.lower() == 'sample name'][0]

        return pd.merge(spreadsheet, fcs_data, 
                        left_on=['Plate#', 'Well position', samples_colname],
                        right_on=['plate', 'well_position', 'sample'], how='left')
    else:
        # print(fcs_data.columns)
        return pd.merge(merged_data, fcs_data, on=['Plate#', 'Well position', 'Sample name'], how='left')



def merge_samples_and_primers(primer_index_file_path, merged_samplesheet_fcs_and_template_sheet_df):
    """
    Merges sample sheet and template sheet data with primer and index information from an Excel file.
    
    Parameters:
    - primer_index_file_path: The path to the Excel file containing primer and index information.
    - merged_samplesheet_fcs_and_template_sheet_df: DataFrame containing merged sample sheet and template sheet data.
    
    Returns:
    - DataFrame with merged information including primer and index data.
    """
    # Read primer and index information from Excel
    primer_index_df = pd.read_excel(primer_index_file_path, sheet_name='Sample primer & index', skiprows=3)
    # Rename 'Sample name' column to 'Sample'
    primer_index_df.rename({'Sample name': 'Sample'}, axis=1, inplace=True)
    
    # Merge the data frames
    merged_df = pd.merge(merged_samplesheet_fcs_and_template_sheet_df, primer_index_df, 
                         on=['Plate#', 'Well position', 'Sample'], 
                         suffixes=('', '_primer'), how='left')
    
    return merged_df


def merge_samples_and_primers(primer_index_file_path, merged_samplesheet_fcs_and_template_sheet_df):
    """
    Merges sample sheet and template sheet data with primer and index information from an Excel file.
    
    Parameters:
    - primer_index_file_path: The path to the Excel file containing primer and index information.
    - merged_samplesheet_fcs_and_template_sheet_df: DataFrame containing merged sample sheet and template sheet data.
    
    Returns:
    - DataFrame with merged information including primer and index data.
    """
    # Read primer and index information from Excel
    primer_index_df = pd.read_excel(primer_index_file_path, sheet_name='Sample primer & index', skiprows=3)
    # Rename 'Sample name' column to 'Sample'
    primer_index_df.rename({'Sample name': 'Sample'}, axis=1, inplace=True)
    
    # Merge the data frames
    merged_df = pd.merge(merged_samplesheet_fcs_and_template_sheet_df, primer_index_df, 
                         on=['Plate#', 'Well position', 'Sample'], 
                         suffixes=('', '_primer'), how='left')
    
    return merged_df